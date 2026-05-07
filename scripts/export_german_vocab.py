#!/usr/bin/env python3
"""Export `German_4000` sheet from the workbook to `src/lib/data/vocab-german-4000.json`.

Requires: pip install openpyxl
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
	from openpyxl import load_workbook
except ImportError:
	print("Install openpyxl: pip install openpyxl", file=sys.stderr)
	sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "lib" / "data" / "vocab-german-4000.json"


def main() -> None:
	if len(sys.argv) < 2:
		print("Usage: export_german_vocab.py <path-to.xlsx>", file=sys.stderr)
		sys.exit(1)
	xlsx = Path(sys.argv[1]).expanduser().resolve()
	if not xlsx.is_file():
		print(f"Not found: {xlsx}", file=sys.stderr)
		sys.exit(1)

	wb = load_workbook(xlsx, read_only=True, data_only=True)
	try:
		ws = wb["German_4000"]
	except KeyError:
		print("Sheet 'German_4000' missing", file=sys.stderr)
		sys.exit(1)

	headers: list[str] | None = None
	rows: list[dict] = []
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		if i == 0:
			headers = [str(h).strip() if h is not None else "" for h in row]
			continue
		rows.append(dict(zip(headers, row)))

	wb.close()

	out: list[dict] = []
	for r in rows:
		out.append(
			{
				"id": r.get("ID"),
				"day": r.get("Day"),
				"week": r.get("Week"),
				"phase": r.get("Phase"),
				"type": r.get("Type"),
				"theme": r.get("Theme"),
				"german": r.get("German entry"),
				"ipa": r.get("Pronunciation (IPA)"),
				"english": r.get("English meaning"),
				"useCase": r.get("Use case"),
			}
		)

	OUT.parent.mkdir(parents=True, exist_ok=True)
	OUT.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
	print(f"Wrote {len(out)} rows -> {OUT}")


if __name__ == "__main__":
	main()
